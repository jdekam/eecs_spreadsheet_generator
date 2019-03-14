import csv
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill

# constants
TIME_RANGE = [
    'AM 8',
    'AM 830',
    'AM 9',
    'AM 930',
    'AM 10',
    'AM 1030',
    'AM 11',
    'AM 1130',
    'PM 12',
    'PM 1230',
    'PM 1',
    'PM 130',
    'PM 2',
    'PM 230',
    'PM 3',
    'PM 330',
    'PM 4',
    'PM 430',
    'PM 5',
    'PM 530',
    'PM 6',
    'PM 630',
    'PM 7',
    'PM 730',
    'PM 8',
    'PM 830',
    'PM 9',
    'PM 930',
    'PM 10'
]
# style info
MED = Side(border_style="medium", color="000000")
BLUE_FILL = PatternFill(start_color = 'ADD8E6', end_color = 'ADD8E6', fill_type = 'solid')
GREEN_FILL = PatternFill(start_color = '98FB98', end_color = '98FB98', fill_type = 'solid')
ORANGE_FILL = PatternFill(start_color = 'FFB732', end_color = 'FFB732', fill_type = 'solid')


# writes time slots into far left row
def write_time(schedule):
    for day in schedule:
        for rowNum in range(0, len(TIME_RANGE) - 1):
            day.cell(row = rowNum + 1, column = 1 ).value = TIME_RANGE[rowNum] + '-' + TIME_RANGE[rowNum+1][3:]


# writes data to .xlsx file
def write_data(day, current_class_num, class_num, sec_num, sec_type, prof_name, start_index, end_index, curr_column):
    # if we are at different class number, move to next open column
    if current_class_num != class_num:
        curr_column += 1

    # else just move to next column if there is overlapping times
    else:
        for rowNum in range(start_index, end_index):
            if day.cell(row = rowNum, column = curr_column).value != None:
                curr_column += 1
                break

    # create data string
    val = class_num + ' ' + sec_type + ' ' + sec_num
    if len(prof_name) > 0:
        val += ' ' + prof_name

    # write and style cell
    for rowNum in range(start_index, end_index):
        day.cell(row = rowNum, column = curr_column).value = val
        # borders
        if rowNum == start_index and rowNum == (end_index - 1):
            day.cell(row = rowNum, column = curr_column).border = Border(right = MED, left = MED, bottom = MED, top = MED)
        elif rowNum == start_index:
            day.cell(row = rowNum, column = curr_column).border = Border(right = MED, left = MED, top = MED)
        elif rowNum == end_index - 1:
            day.cell(row = rowNum, column = curr_column).border = Border(right = MED, left = MED, bottom = MED)
        else:
            day.cell(row = rowNum, column = curr_column).border = Border(right = MED, left = MED)

        # colors
        if sec_type == 'LEC':
            day.cell(row = rowNum, column = curr_column).fill = BLUE_FILL
        elif sec_type == 'LAB':
            day.cell(row = rowNum, column = curr_column).fill = GREEN_FILL
        elif sec_type == 'DIS':
            day.cell(row = rowNum, column = curr_column).fill = ORANGE_FILL

    # return column
    return curr_column


def main():
    # open csv file
    src_file = open('resources/src.csv', 'r+')
    src_data = csv.reader(src_file)
    src = [row for row in src_data]

    # create new excel workbook / sheets
    schedule = Workbook()
    monday = schedule.active
    monday.title = 'Monday'
    tuesday = schedule.create_sheet('Tuesday')
    wednesday = schedule.create_sheet('Wednesday')
    thursday = schedule.create_sheet('Thursday')
    friday = schedule.create_sheet('Friday')

    # write time slots on far left
    write_time(schedule)

    # initialize values of col by day
    m_col = 1
    t_col = 1
    w_col = 1
    th_col = 1
    f_col = 1

    # initialize list of current classes to lowest one offered by day
    lc = [ ]
    for day in range(3, 8):
        for row in src:
            if not (row[day] == ""):
                lc.append(row[day])
                break

    # loop through course schedule
    for row in src:
        # pull data from CSV file
        class_num = row[0]
        sec_num = row[1]
        sec_type = row[2]

        # list of days where class takes place
        classOccurs = [ ]
        for i in range(3, 8):
            classOccurs.append(not (row[i] == ''))

        # calculate start and end times
        start = row[8][-2:] + ' ' + row[8].split('-')[0]
        # need this check for classes that start in an AM time and end in a PM time
        if start == 'PM 1130' or start == 'PM 11' or start == 'PM 1030' or start == 'PM 10' or start == 'PM 930' or start == 'PM 9':
            start = 'AM ' + start[3:]
        end = row[8][-2:] + ' '+ row[8].split('-')[1][:-2]
        # calculate beginning and ending indexes to write to (incremented by 1 since pyxl is 1-indexed)
        start_index = TIME_RANGE.index(start) + 1
        end_index = TIME_RANGE.index(end) + 1

        # if there is a professer name saved read that in
        prof_name = row[9]
            
        # write data to appropriate sheets
        if classOccurs[0]:
            m_col = write_data(monday, lc[0], class_num, sec_num, sec_type, prof_name, start_index, end_index, m_col)
            if lc[0] != class_num:
                lc[0] = class_num
        if classOccurs[1]:
            t_col = write_data(tuesday, lc[1], class_num, sec_num, sec_type, prof_name, start_index, end_index, t_col)
            if lc[1] != class_num:
                lc[1] = class_num
        if classOccurs[2]:
            w_col = write_data(wednesday, lc[2], class_num, sec_num, sec_type, prof_name, start_index, end_index, w_col)
            if lc[2] != class_num:
                lc[2] = class_num
        if classOccurs[3]:
            th_col = write_data(thursday, lc[3], class_num, sec_num, sec_type, prof_name, start_index, end_index, th_col)
            if lc[3] != class_num:
                lc[3] = class_num
        if classOccurs[4]:
            f_col = write_data(friday, lc[4], class_num, sec_num, sec_type, prof_name, start_index, end_index, f_col)
            if lc[4] != class_num:
                lc[4] = class_num

    # save finished workbook
    schedule.save('schedule.xlsx')

if __name__ == '__main__':
    main()